"""
Comparison Report & Final Excel Generator
=========================================
Reads `kanyakumari_scrape_audit_v2.json` and produces:
1. Updated Excel Dataset: `kanyakumari_colleges_validated_v2_updated.xlsx` & `kanyakumari_colleges_validated_v2.xlsx`
2. Side-by-Side Comparison Report: `kanyakumari_colleges_comparison_report.xlsx` & `kanyakumari_colleges_comparison_report.csv`
"""

import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AUDIT_JSON_PATH = "D:/LINKEDSTORY/Projects/kanyakumari_scrape_audit_v2.json"
UPDATED_EXCEL_PATH = "D:/LINKEDSTORY/Projects/kanyakumari_colleges_validated_v2_updated.xlsx"
DIRECT_EXCEL_PATH = "D:/LINKEDSTORY/Projects/kanyakumari_colleges_validated_v2.xlsx"
COMPARISON_EXCEL_PATH = "D:/LINKEDSTORY/Projects/kanyakumari_colleges_comparison_report.xlsx"
COMPARISON_CSV_PATH = "D:/LINKEDSTORY/Projects/kanyakumari_colleges_comparison_report.csv"

def generate_reports():
    print(f"[*] Reading audit JSON: {AUDIT_JSON_PATH}")
    with open(AUDIT_JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    # -------------------------------------------------------------------------
    # DELIVERABLE 1: Updated Excel Directory File
    # -------------------------------------------------------------------------
    for save_path in [UPDATED_EXCEL_PATH, DIRECT_EXCEL_PATH]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Colleges_Data"

        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        font_data = Font(name="Calibri", size=10)
        fill_alt = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        headers = [
            "S.No", "Category", "College Name", "Location", "Website",
            "Principal Name", "Principal Email", "Principal Phone", "General Contact",
            "Courses Offered", "Departments", "Total Student Strength (Approx)",
            "Department-Wise Student Count Breakdown"
        ]

        ws.row_dimensions[1].height = 28
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row_idx, item in enumerate(data, 2):
            row_fill = fill_alt if row_idx % 2 == 0 else fill_white
            ws.row_dimensions[row_idx].height = 36

            vals = [
                item["sno"],
                item["category"],
                item["college_name"],
                item["location"],
                item["verified_website"],
                item["verified_principal"],
                item["verified_email"],
                item["verified_phone"],
                item["verified_general"],
                item["courses"],
                item["departments"],
                item["students"],
                item["breakdown"]
            ]

            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_data
                cell.fill = row_fill
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        col_widths = {1: 8, 2: 24, 3: 38, 4: 30, 5: 32, 6: 34, 7: 34, 8: 26, 9: 34, 10: 45, 11: 50, 12: 28, 13: 60}
        for col_idx, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.auto_filter.ref = f"A1:M{len(data)+1}"
        wb.save(save_path)
        print(f"[+] Saved updated Excel directory to: {save_path}")

    # -------------------------------------------------------------------------
    # DELIVERABLE 2: Side-by-Side Comparison Report (Excel & CSV)
    # -------------------------------------------------------------------------
    comparison_rows = []
    
    fields_to_compare = [
        ("Official Website", "orig_website", "verified_website"),
        ("Principal / Dean Name", "orig_principal", "verified_principal"),
        ("Principal Contact Email", "orig_email", "verified_email"),
        ("Principal Contact Phone", "orig_phone", "verified_phone"),
        ("General Contact Info", "orig_general", "verified_general")
    ]

    for item in data:
        for field_label, orig_key, ver_key in fields_to_compare:
            orig_val = item.get(orig_key, "")
            ver_val = item.get(ver_key, "")

            # Determine field status
            if orig_val == ver_val:
                status = "Verified (No Change)"
            elif ("Not Available" in orig_val or not orig_val) and ("Not Available" not in ver_val and ver_val):
                status = "Updated (Filled Missing Info)"
            elif orig_val != ver_val:
                status = "Corrected (Replaced Inaccurate Entry)"
            else:
                status = "Unverified"

            comparison_rows.append({
                "S.No": item["sno"],
                "Category": item["category"],
                "College Name": item["college_name"],
                "Field Evaluated": field_label,
                "Original Value (Inaccurate/Missing)": orig_val,
                "Scraped & Verified Value": ver_val,
                "Field Status": status,
                "Verification Level": item["verification_status"],
                "Primary Source Citation": item["primary_source"],
                "Secondary Source Citation": item["secondary_source"],
                "Discrepancy & Verification Notes": item["notes"]
            })

    # Save CSV comparison report
    csv_headers = list(comparison_rows[0].keys())
    with open(COMPARISON_CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=csv_headers)
        writer.writeheader()
        writer.writerows(comparison_rows)
    print(f"[+] Saved CSV comparison report to: {COMPARISON_CSV_PATH}")

    # Save Excel comparison report workbook
    wb_comp = openpyxl.Workbook()
    ws_comp = wb_comp.active
    ws_comp.title = "Side_By_Side_Comparison"

    font_header_comp = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header_comp = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    font_data_comp = Font(name="Calibri", size=10)
    fill_corrected = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light yellow
    fill_updated = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Light green
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws_comp.row_dimensions[1].height = 28
    for col_idx, h in enumerate(csv_headers, 1):
        cell = ws_comp.cell(row=1, column=col_idx, value=h)
        cell.font = font_header_comp
        cell.fill = fill_header_comp
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for r_idx, row_dict in enumerate(comparison_rows, 2):
        ws_comp.row_dimensions[r_idx].height = 32
        status = row_dict["Field Status"]
        
        row_fill = fill_white
        if "Corrected" in status:
            row_fill = fill_corrected
        elif "Updated" in status:
            row_fill = fill_updated

        for col_idx, h_name in enumerate(csv_headers, 1):
            cell = ws_comp.cell(row=r_idx, column=col_idx, value=row_dict[h_name])
            cell.font = font_data_comp
            cell.fill = row_fill
            cell.border = thin_border
            if col_idx in [1]:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    comp_col_widths = {
        1: 8, 2: 24, 3: 38, 4: 24, 5: 35, 6: 35, 7: 32, 8: 38, 9: 42, 10: 45, 11: 50
    }
    for col_idx, w in comp_col_widths.items():
        ws_comp.column_dimensions[get_column_letter(col_idx)].width = w

    ws_comp.auto_filter.ref = f"A1:K{len(comparison_rows)+1}"
    wb_comp.save(COMPARISON_EXCEL_PATH)
    print(f"[+] Saved Excel comparison report to: {COMPARISON_EXCEL_PATH}")

if __name__ == "__main__":
    generate_reports()
