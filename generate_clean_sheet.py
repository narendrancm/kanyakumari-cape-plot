"""
Generate clean, plain, natural spreadsheet without colorful AI-like styling.
Looks like a standard student/developer web scraping output.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import the dataset from scrape_kanyakumari_colleges
from scrape_kanyakumari_colleges import COLLEGES_DATA

def create_simple_excel(filename="d:/LINKEDSTORY/Projects/kanyakumari_colleges.xlsx"):
    print(f"Creating natural plain Excel file: {filename}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colleges_Data"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True

    # Standard clean fonts - plain black, no flashy colors
    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    data_font = Font(name="Calibri", size=11, color="000000")
    
    # Optional simple light gray header fill (standard Excel look) or plain white
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Standard thin borders
    border_style = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    headers = [
        "S.No",
        "Category",
        "College Name",
        "Location",
        "Website",
        "Principal Name",
        "Principal Email",
        "Principal Phone",
        "General Contact",
        "Courses Offered",
        "Departments",
        "Total Students (Approx)",
        "Department-wise Student Strength"
    ]

    # Row 1: Headers
    ws.row_dimensions[1].height = 24
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")
        cell.border = border_style

    # Data Rows
    sno = 1
    for r_idx, item in enumerate(COLLEGES_DATA, start=2):
        ws.row_dimensions[r_idx].height = 36
        
        # Clean category name without numbering prefix
        category_clean = item["category"].split(". ")[-1]
        
        row_vals = [
            sno,
            category_clean,
            item["name"],
            item["location"],
            item["website"],
            item["principal_head"],
            item["principal_email"],
            item["principal_phone"],
            item["general_contact"],
            item["courses_offered"],
            item["departments"],
            item.get("total_students", "N/A"),
            item.get("dept_students_breakdown", "N/A")
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_style
            
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        sno += 1

    # Simple column widths
    col_widths = {
        1: 8,    # S.No
        2: 24,   # Category
        3: 38,   # College Name
        4: 30,   # Location
        5: 28,   # Website
        6: 28,   # Principal Name
        7: 32,   # Principal Email
        8: 24,   # Principal Phone
        9: 28,   # General Contact
        10: 42,  # Courses Offered
        11: 48,  # Departments
        12: 26,  # Total Students
        13: 55   # Dept Students
    }
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Add standard auto-filter
    ws.auto_filter.ref = f"A1:M{len(COLLEGES_DATA) + 1}"

    # Save
    try:
        wb.save(filename)
        print(f"[OK] Clean sheet saved to {filename}")
    except PermissionError:
        alt_filename = "d:/LINKEDSTORY/Projects/kanyakumari_colleges_simple.xlsx"
        wb.save(alt_filename)
        print(f"[OK] Saved to alternative file: {alt_filename}")

if __name__ == "__main__":
    create_simple_excel("d:/LINKEDSTORY/Projects/kanyakumari_colleges_simple.xlsx")
    create_simple_excel("d:/LINKEDSTORY/Projects/kanyakumari_colleges.xlsx")
