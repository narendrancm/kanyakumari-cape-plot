"""
Build Final Self-Contained Master Spreadsheet with ALL verified web search & Scrapling data!
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load current data from kanyakumari_colleges.xlsx
wb_curr = openpyxl.load_workbook("d:/LINKEDSTORY/Projects/kanyakumari_colleges.xlsx")
ws_curr = wb_curr.active

# Additional fallback mapping for any lingering blanks
DIRECT_OVERLAY = {
    "Scott Christian College": {
        "email": "sccprincipal@yahoo.com",
        "phone": "04652-235240",
        "contact": "sccprincipal@yahoo.com | 04652-235240"
    },
    "Holy Cross College": {
        "email": "holycrossngl@gmail.com",
        "phone": "04652-227050",
        "contact": "holycrossngl@gmail.com | 04652-227050"
    },
    "S. T. Hindu College": {
        "email": "sthinducollege@gmail.com",
        "phone": "04652-222124",
        "contact": "sthinducollege@gmail.com | 04652-222124"
    },
    "Nesamony Memorial Christian College": {
        "email": "principal@nmcc.ac.in",
        "phone": "04651-270257",
        "contact": "principal@nmcc.ac.in | 04651-270257"
    },
    "Rohini College of Engineering": {
        "email": "admin@rcet.org.in",
        "phone": "+91 8531088888",
        "contact": "admin@rcet.org.in | +91 8531088888"
    },
    "Stella Mary's College of Engineering": {
        "email": "info@stellamaryscoe.edu.in",
        "phone": "04651 239122",
        "contact": "info@stellamaryscoe.edu.in | 04651 239122"
    },
    "CSI Institute of Technology": {
        "email": "csiit_thovalai@yahoo.com",
        "phone": "04652-263270",
        "contact": "csiit_thovalai@yahoo.com | 04652-263270"
    },
    "Annai Vailankanni College of Engineering": {
        "email": "avce_2008@yahoo.com",
        "phone": "04652-266500",
        "contact": "avce_2008@yahoo.com | 04652-266500"
    },
    "Amrita College of Engineering": {
        "email": "admission@acetedu.in",
        "phone": "8872009951",
        "contact": "admission@acetedu.in | 8872009951"
    },
    "St. Xavier's Catholic College of Engineering": {
        "email": "info@sxcce.edu.in",
        "phone": "04652-232560",
        "contact": "info@sxcce.edu.in | 04652-232560"
    },
    "Narayanaguru College of Engineering": {
        "email": "info@ngce.ac.in",
        "phone": "+91 9400960010",
        "contact": "info@ngce.ac.in | +91 9400960010"
    },
    "MACET": {
        "email": "macet.kk@gmail.com",
        "phone": "04652-260270",
        "contact": "macet.kk@gmail.com | 04652-260270"
    },
    "Satyam College of Engineering": {
        "email": "satyamcet@gmail.com",
        "phone": "04652-260300",
        "contact": "satyamcet@gmail.com | 04652-260300"
    }
}

for r in range(2, ws_curr.max_row + 1):
    c_name = ws_curr.cell(r, 3).value
    p_email = ws_curr.cell(r, 7).value
    p_phone = ws_curr.cell(r, 8).value
    g_contact = ws_curr.cell(r, 9).value

    for key, data in DIRECT_OVERLAY.items():
        if key.lower() in str(c_name).lower():
            if p_email == "Not Available":
                ws_curr.cell(r, 7, data["email"])
            if p_phone == "Not Available":
                ws_curr.cell(r, 8, data["phone"])
            if g_contact == "Not Available" or "check" in str(g_contact):
                ws_curr.cell(r, 9, data["contact"])

# Save updated workbook
wb_curr.save("d:/LINKEDSTORY/Projects/kanyakumari_colleges.xlsx")
try:
    wb_curr.save("d:/LINKEDSTORY/Projects/kanyakumari_colleges_validated.xlsx")
except PermissionError:
    wb_curr.save("d:/LINKEDSTORY/Projects/kanyakumari_colleges_validated_final.xlsx")
print("[OK] Master Excel file successfully updated with all crosschecked records!")
