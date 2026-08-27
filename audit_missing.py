import openpyxl

wb = openpyxl.load_workbook("d:/LINKEDSTORY/Projects/kanyakumari_colleges.xlsx")
ws = wb.active

missing_colleges = []

for r in range(2, ws.max_row + 1):
    c_name = ws.cell(r, 3).value
    cat = ws.cell(r, 2).value
    web = ws.cell(r, 5).value
    p_name = ws.cell(r, 6).value
    p_email = ws.cell(r, 7).value
    p_phone = ws.cell(r, 8).value
    gen_contact = ws.cell(r, 9).value

    if (web == "Not Available" or p_email == "Not Available" or p_phone == "Not Available" or p_name == "Not Available"):
        missing_colleges.append({
            "row": r,
            "category": cat,
            "name": c_name,
            "website": web,
            "principal_name": p_name,
            "principal_email": p_email,
            "principal_phone": p_phone,
            "general_contact": gen_contact
        })

print(f"Total Colleges Needing Deeper Lookup: {len(missing_colleges)} / {ws.max_row - 1}")
for item in missing_colleges:
    print(f"Row {item['row']:2d} [{item['category']}] {item['name']}")
    print(f"   Web: {item['website']} | Email: {item['principal_email']} | Phone: {item['principal_phone']}")
